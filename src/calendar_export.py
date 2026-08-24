"""
Regenerates the 90-day content calendar as an .xlsx workbook (two sheets:
"90-Day Calendar" and "Milestones"). This is the same generator used to
produce the original calendar delivered earlier -- kept here so the
calendar can be edited/extended in code (e.g. adding weeks 14+) and
re-exported, instead of hand-editing the spreadsheet.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.config import config

HEADER_FILL = "1F3864"
THEME_FILL = "DCE6F1"

DEFAULT_WEEKS = [
    (1, "Launch / Counting", [
        ("Mon", "New Episode", "Ep1: Counting Carrots (1-10)"),
        ("Wed", "New Episode", "Ep2: Counting Flowers"),
        ("Fri", "New Episode", "Ep3: Counting Clouds"),
        ("Sun", "Compilation", '"Bounce Counts to 10!" (3-ep compilation)'),
    ]),
    (2, "Counting (cont.)", [
        ("Mon", "New Episode", "Ep4: Counting Birds"),
        ("Wed", "New Episode", "Ep5: Counting Leaves"),
        ("Fri", "New Episode", "Ep6: Counting Friends"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (3, "Colors", [
        ("Mon", "New Episode", "Ep7: Red & Orange"),
        ("Wed", "New Episode", "Ep8: Yellow & Green"),
        ("Fri", "New Episode", "Ep9: Blue & Purple"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (4, "Colors (cont.)", [
        ("Mon", "New Episode", "Ep10: Rainbow Song"),
        ("Wed", "New Episode", "Ep11: Color Hunt in the Meadow"),
        ("Fri", "New Episode", "Ep12: Mixing Colors (paint play)"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (5, "Shapes", [
        ("Mon", "New Episode", "Ep13: Circle & Square"),
        ("Wed", "New Episode", "Ep14: Triangle & Star"),
        ("Fri", "New Episode", "Ep15: Shape Hunt Outdoors"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (6, "Animal Sounds", [
        ("Mon", "New Episode", "Ep16: Farm Animals"),
        ("Wed", "New Episode", "Ep17: Forest Animals"),
        ("Fri", "New Episode", "Ep18: Ocean Animals"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (7, "Alphabet - Part 1", [
        ("Mon", "New Episode", "Ep19: A-E"),
        ("Wed", "New Episode", "Ep20: F-J"),
        ("Fri", "New Episode", "Ep21: K-O"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (8, "Alphabet - Part 2", [
        ("Mon", "New Episode", "Ep22: P-T"),
        ("Wed", "New Episode", "Ep23: U-Z"),
        ("Fri", "New Episode", "Ep24: Alphabet Song (full)"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (9, "Emotions & Manners", [
        ("Mon", "New Episode", "Ep25: Happy / Sad / Excited"),
        ("Wed", "New Episode", "Ep26: Sharing & Kindness"),
        ("Fri", "New Episode", "Ep27: Please & Thank You"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (10, "Weather & Seasons", [
        ("Mon", "New Episode", "Ep28: Sunny Day"),
        ("Wed", "New Episode", "Ep29: Rainy Day"),
        ("Fri", "New Episode", "Ep30: Four Seasons Song"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (11, "Bedtime & Calm", [
        ("Mon", "New Episode", "Ep31: Bounce's Bedtime Song"),
        ("Wed", "New Episode", "Ep32: Counting Sheep (original)"),
        ("Fri", "New Episode", "Ep33: Goodnight Meadow"),
        ("Sun", "Compilation", "Weekly compilation"),
    ]),
    (12, "Fan-Favorite Remix", [
        ("Mon", "New Episode", "Ep34: (repeat top-performing topic, new twist)"),
        ("Wed", "New Episode", "Ep35: (repeat top-performing topic, new twist)"),
        ("Fri", "New Episode", "Ep36: (repeat top-performing topic, new twist)"),
        ("Sun", "Compilation", "Big weekly compilation"),
    ]),
    (13, "Wrap-Up / Seasonal Tie-In", [
        ("Mon", "New Episode", "Ep37"),
        ("Wed", "New Episode", "Ep38"),
        ("Fri", "New Episode", "Ep39 (finale / teaser for next arc)"),
        ("Sun", "Compilation", '"Best of First 90 Days" compilation'),
    ]),
]

DEFAULT_MILESTONES = [
    ("Day 1", "Publish channel trailer (30s intro to Bounce), set channel art/branding, mark ALL uploads 'Made for Kids', disable personalized ads and comments per COPPA requirements."),
    ("Day 7", "First weekly compilation goes live. Check audience retention graph on Week 1 originals; adjust pacing for Week 2 scripts if drop-off is early."),
    ("Day 14", "Review Week 1-2 analytics (avg view duration, replay rate, top-performing episode) before locking in Week 3+ themes."),
    ("Day 30", "Track YouTube Partner Program eligibility (watch-hours / subscribers). Identify best-performing theme so far to inform the Week 12 remix slot."),
    ("Day 45", "Begin outreach for brand / toy-company sponsorship conversations (lead times are long -- start before revenue is urgently needed)."),
    ("Day 60", "Consider first dub / localization (e.g. Spanish) of the top 5 performing episodes to test international reach."),
    ("Day 90", "Full 90-day analytics review. Decide Phase 2 content pillars: new character arc, merch test, or licensing outreach."),
]


def generate_calendar_workbook(
    weeks=None,
    milestones=None,
    out_path: Optional[str] = None,
) -> Path:
    weeks = weeks or DEFAULT_WEEKS
    milestones = milestones or DEFAULT_MILESTONES

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "90-Day Calendar"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    theme_font = Font(name="Arial", bold=True, size=11)
    theme_fill = PatternFill(start_color=THEME_FILL, end_color=THEME_FILL, fill_type="solid")
    normal_font = Font(name="Arial", size=11)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Week", "Theme", "Day", "Upload Type", "Episode Title", "Notes / Status"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row = 2
    for week_num, theme, days in weeks:
        first_row = row
        for day, upload_type, title in days:
            values = [week_num, theme, day, upload_type, title, ""]
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1
        ws.merge_cells(start_row=first_row, start_column=1, end_row=row - 1, end_column=1)
        ws.merge_cells(start_row=first_row, start_column=2, end_row=row - 1, end_column=2)
        for c in (1, 2):
            cell = ws.cell(row=first_row, column=c)
            cell.font = theme_font
            cell.fill = theme_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, width in {"A": 7, "B": 24, "C": 8, "D": 14, "E": 42, "F": 30}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    ms = wb.create_sheet("Milestones")
    ms.append(["Day", "Milestone / Action"])
    for c in range(1, 3):
        cell = ms.cell(row=1, column=c)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, (day, action) in enumerate(milestones, start=2):
        ms.cell(row=r, column=1, value=day).font = Font(name="Arial", bold=True, size=11)
        ms.cell(row=r, column=2, value=action).font = normal_font
        for c in range(1, 3):
            ms.cell(row=r, column=c).border = border
            ms.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    ms.column_dimensions["A"].width = 10
    ms.column_dimensions["B"].width = 100
    ms.freeze_panes = "A2"
    ms.row_dimensions[1].height = 20

    out_dir = Path(config.OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path_p = Path(out_path) if out_path else out_dir / "Bounce_90_Day_Content_Calendar.xlsx"
    wb.save(out_path_p)
    print(f"Saved calendar to {out_path_p}")
    return out_path_p
